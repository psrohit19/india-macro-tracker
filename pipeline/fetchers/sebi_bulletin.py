"""
SEBI monthly-bulletin fetcher — LIVE. Populates: ipo_qip (monthly, ₹ crore).

ipo_qip = IPO (Total: Main Board + SME, incl. OFS) + QIP/IPP amounts, from
the SEBI Bulletin's "Annexure Tables" workbook:

  * Sheet "6"  — Table 6: Consolidated Resource Mobilisation… row
                 "C. IPO (Total) [A+B]"; header row has one column pair
                 (count, ₹ cr amount) per month of the fiscal year to date.
                 Month headers are datetimes (month-start OR month-end,
                 varies) or strings like "Jun-25 (revised)".
  * Sheet "12" — Table 12: …through QIPs; one row per month (datetime in
                 col 0), total amount in col 10.

The workbook URL changes every edition and the SEBI publications listing is not
statically crawlable, so ANNEXURE_URL is pinned to the newest known edition.
TODO when a new bulletin lands (~1 month lag; May-2026 edition next):
update ANNEXURE_URL from the bulletin page
https://www.sebi.gov.in/reports-and-statistics/publications/<mon>-<yyyy>/…
("[Annexure Tables]" link). Cell values are cached formula results, so the
workbook must be loaded with data_only=True.

FY2024-25 months are PINNED below — extracted once from the March-2025
edition (commondocs/mar-2025/SEBI Bulletin Excel_p.xlsx), whose layout
differs (Financial/Non-financial split, Tables 5 & 10). Mar-2025 itself was
still empty in that edition and remains a gap.
"""
import datetime as dt
import io
import json
import re
from pathlib import Path

import openpyxl
import requests

SERIES_IDS = ["ipo_qip"]
ANNEXURE_URL = ("https://www.sebi.gov.in/sebi_data/commondocs/apr-2026/"
                "Annexure%20Tables-30.04.2026_p.xlsx")
H = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/126.0"}
LIVE = Path(__file__).parent.parent.parent / "data" / "live"

MONTHS = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}

# FY2024-25, ₹ cr: IPO(Total) + QIP totals from the Mar-2025 edition
# (Tables 5 & 10; IPO = Financial + Non-financial amounts).
PINNED = {
    "2024-04-01": 5727.40 + 11471.73, "2024-05-01": 10133.35 + 3040.00,
    "2024-06-01": 2520.76 + 2774.84,  "2024-07-01": 6060.24 + 13699.10,
    "2024-08-01": 15469.32 + 12281.60, "2024-09-01": 16212.78 + 21482.76,
    "2024-10-01": 34982.99 + 15589.00, "2024-11-01": 35848.64 + 11190.05,
    "2024-12-01": 27031.19 + 34739.49, "2025-01-01": 3065.60 + 3960.65,
    "2025-02-01": 14745.33 + 0.0,
}


def _month_key(v):
    """Header cell -> 'YYYY-MM-01' or None. Handles datetime / 'Jun-25 (…)'."""
    if isinstance(v, dt.datetime):
        return f"{v.year:04d}-{v.month:02d}-01"
    if isinstance(v, str):
        m = re.match(r"([A-Z][a-z]{2})-(\d{2})", v.strip())
        if m and m.group(1) in MONTHS:
            return f"{2000 + int(m.group(2)):04d}-{MONTHS[m.group(1)]:02d}-01"
    return None


def fetch():
    r = requests.get(ANNEXURE_URL, headers=H, timeout=120)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True,
                                data_only=True)

    ws = wb["6"]
    rows = list(ws.iter_rows(values_only=True, max_row=45))
    hdr = rows[1]
    ipo_row = next(r_ for r_ in rows
                   if r_[0] and str(r_[0]).startswith("C. IPO (Total)"))
    ipo = {}
    for i, cell in enumerate(hdr):
        key = _month_key(cell)
        if key and ipo_row[i + 1] is not None:      # amount sits next to count
            ipo[key] = float(ipo_row[i + 1])

    qip = {}
    for row in wb["12"].iter_rows(values_only=True):
        key = _month_key(row[0])
        if key and row[10] is not None:
            qip[key] = float(row[10])

    obs = dict(PINNED)
    for k in ipo:
        obs[k] = round(ipo[k] + qip.get(k, 0.0), 1)
    obs = {k: round(v, 1) for k, v in obs.items()}

    LIVE.mkdir(parents=True, exist_ok=True)
    (LIVE / "ipo_qip.json").write_text(
        json.dumps({"freq": "M", "obs": sorted(obs.items())}))
    last = sorted(obs)[-1]
    print(f"  ipo_qip: {len(obs)} obs, latest {last} = {obs[last]:,.0f} cr")
    return obs


if __name__ == "__main__":
    fetch()
